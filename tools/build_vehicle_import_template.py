from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "車籍資料匯入範本.xlsx"

HEADERS = [
    "車牌號碼",
    "車主姓名",
    "手機",
    "車種",
    "出廠年月",
    "初次領牌日",
    "上次檢驗日",
    "下次檢驗日",
    "通知同意",
    "通知方式",
    "Email",
    "地址",
    "身分證末四碼",
    "車身號碼末四碼",
    "燃料種類",
    "排氣量",
    "備註",
]

SAMPLES = [
    ["ABC-1234", "王先生", "0912-345-678", "自用小客車", "2014-05", "2014-06-15", "2026-06-23", "", "同意", "電話", "", "", "", "", "汽油", "1800", "電話預約補登"],
    ["KLD-8899", "林小姐", "0988-222-123", "自用小客車", "2019-11", "2019-12-02", "2026-06-23", "", "同意", "LINE", "", "", "", "", "油電", "1600", ""],
    ["TX-5678", "陳先生", "0975-111-888", "自用小貨車", "2012-03", "2012-04-10", "2026-06-23", "", "同意", "簡訊", "", "", "", "", "柴油", "2400", "十年以上車輛"],
]

FIELD_GUIDE = [
    ["欄位", "必要性", "格式/範例", "用途", "空白處理"],
    ["車牌號碼", "必填", "ABC-1234", "車輛唯一識別與搜尋", "空白列不匯入"],
    ["車主姓名", "必填", "王先生", "聯絡與資料顯示", "空白時以未填寫顯示"],
    ["手機", "必填", "0912-345-678", "通知、搜尋與去重", "空白時仍可匯入但列為待補"],
    ["車種", "建議", "自用小客車", "檢驗規則與分類", "預設為自用小客車"],
    ["出廠年月", "建議", "2014-05", "判斷車齡與檢驗週期", "空白時無法自動判斷週期"],
    ["初次領牌日", "選填", "2014-06-15", "補充車籍資料", "可空白"],
    ["上次檢驗日", "建議", "2026-06-23", "推算下次提醒", "空白時以匯入日或人工補登"],
    ["下次檢驗日", "選填", "2026-12-23", "若車廠已有資料可直接指定", "空白時由系統依車齡推算"],
    ["通知同意", "建議", "同意/未同意", "決定是否列入提醒名單", "預設未同意"],
    ["通知方式", "選填", "電話/LINE/簡訊/Email", "後續聯絡偏好", "可空白"],
    ["Email", "選填", "owner@example.com", "補充通知管道", "可空白"],
    ["地址", "選填", "完整地址", "車廠內部備查", "可空白"],
    ["身分證末四碼", "選填", "1234", "輔助核對，不建議存完整身分證", "可空白"],
    ["車身號碼末四碼", "選填", "5678", "輔助核對，不建議存完整車身號碼", "可空白"],
    ["燃料種類", "選填", "汽油/柴油/油電/電動", "車輛分類", "可空白"],
    ["排氣量", "選填", "1800", "車輛分類", "可空白"],
    ["備註", "選填", "文字", "現場補充說明", "可空白"],
]

OPTIONS = [
    ["車種", "通知同意", "通知方式", "燃料種類"],
    ["自用小客車", "同意", "電話", "汽油"],
    ["自用小貨車", "未同意", "LINE", "柴油"],
    ["營業車", "", "簡訊", "油電"],
    ["機車", "", "Email", "電動"],
    ["其他", "", "", "其他"],
]


def style_header(ws, columns):
    fill = PatternFill("solid", fgColor="236F78")
    for cell in ws[1][:columns]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_rows(ws, rows):
    for row in rows:
        ws.append(row)


wb = Workbook()
data = wb.active
data.title = "車籍資料匯入"
guide = wb.create_sheet("欄位說明")
options = wb.create_sheet("下拉選項")

write_rows(data, [HEADERS, *SAMPLES])
write_rows(guide, FIELD_GUIDE)
write_rows(options, OPTIONS)

for ws, count in [(data, len(HEADERS)), (guide, 5), (options, 4)]:
    style_header(ws, count)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

widths = {
    "A": 16,
    "B": 14,
    "C": 16,
    "D": 14,
    "E": 12,
    "F": 14,
    "G": 14,
    "H": 14,
    "I": 12,
    "J": 12,
    "K": 22,
    "L": 28,
    "M": 14,
    "N": 16,
    "O": 12,
    "P": 10,
    "Q": 28,
}
for col, width in widths.items():
    data.column_dimensions[col].width = width
for col in ["A", "B", "C", "D", "E"]:
    guide.column_dimensions[col].width = 24 if col not in ["D", "E"] else 34
for col in ["A", "B", "C", "D"]:
    options.column_dimensions[col].width = 16

data.auto_filter.ref = "A1:Q1"
guide.auto_filter.ref = "A1:E1"

validations = {
    "D2:D500": '"自用小客車,自用小貨車,營業車,機車,其他"',
    "I2:I500": '"同意,未同意"',
    "J2:J500": '"電話,LINE,簡訊,Email"',
    "O2:O500": '"汽油,柴油,油電,電動,其他"',
}
for target, formula in validations.items():
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    data.add_data_validation(dv)
    dv.add(target)

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
