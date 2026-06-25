"""產生 Google Sheet 匯入用 Excel 範本"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F6F78", end_color="1F6F78", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D7DEE5"),
    right=Side(style="thin", color="D7DEE5"),
    top=Side(style="thin", color="D7DEE5"),
    bottom=Side(style="thin", color="D7DEE5"),
)

# ── 各週工作表欄位 ──
weekly_headers = ["組別", "組員", "任務完成項目", "遇到的問題", "自我評量(1-5)", "教師備註"]
weekly_widths = [12, 28, 40, 36, 14, 30]

sample_groups = [
    ("第1組", "王小明, 李小華, 張大中"),
    ("第2組", "陳小美, 林小君"),
    ("第3組", "黃小豪, 吳小安"),
    ("第4組", "劉小華, 周小倫"),
    ("第5組", "林小宏, 許小庭"),
]

# ── 建立各週工作表 ──
for w in range(1, 21):
    ws = wb.create_sheet(title=f"W{w}") if w > 1 else wb.active
    ws.title = f"W{w}"
    for i, h in enumerate(weekly_headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    for ci, ww in enumerate(weekly_widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = ww
    ws.freeze_panes = "A2"
    for ri, (group, members) in enumerate(sample_groups, 2):
        ws.cell(row=ri, column=1, value=group).border = thin_border
        ws.cell(row=ri, column=2, value=members).border = thin_border
        for ci in range(3, 7):
            ws.cell(row=ri, column=ci).border = thin_border

# ── 總覽工作表 ──
overview = wb.create_sheet(title="總覽", index=0)
overview_headers = ["組別"]
for w in range(1, 21):
    overview_headers.append(f"W{w}完成")
    overview_headers.append(f"W{w}評量")
overview_headers.append("備註")
for i, h in enumerate(overview_headers, 1):
    cell = overview.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
overview.column_dimensions["A"].width = 12
for ci in range(2, len(overview_headers) + 1):
    overview.column_dimensions[chr(64 + ci) if ci <= 26 else "A"].width = 10
overview.freeze_panes = "B2"
for ri, (group, _) in enumerate(sample_groups, 2):
    overview.cell(row=ri, column=1, value=group).border = thin_border
    for ci in range(2, len(overview_headers) + 1):
        overview.cell(row=ri, column=ci).border = thin_border

output = "J:\\我的雲端硬碟\\GPTcodex資料夾\\20260622驗車場專案\\data\\學習歷程範本.xlsx"
wb.save(output)
print(f"已產生：{output}")
